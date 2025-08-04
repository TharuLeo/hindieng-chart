from flask import Flask, render_template, request, send_file
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import datetime
import re

app = Flask(__name__)

# Load TRAIN TIMING sheet once at startup
train_timing_df = pd.read_excel("train-timing.xlsx", sheet_name="TRAIN TIMING")
station_order = train_timing_df["Station Code"].tolist()
hindi_station_map = dict(zip(train_timing_df["Station Code"], train_timing_df["Hindi Name"]))

# Column headers from TRAIN TIMING sheet
eng_headers = train_timing_df.columns[6:9].tolist()
hin_headers = train_timing_df.iloc[0, 6:9].tolist()

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        raw_data = request.form.get("coach_data")
        lines = raw_data.strip().split("\n")
        grouped = {}
        train_number = ""
        train_date = ""

        for line in lines:
            parts = line.strip().split("|")
            if len(parts) >= 5:
                coach = parts[1].strip()
                code = parts[2].strip()
                berth = parts[3].strip()
                time = parts[4].strip()

                if coach not in grouped:
                    grouped[coach] = []

                grouped[coach].append([code, berth, time])

                if not train_number and re.search(r'\d{5}', line):
                    train_number = re.search(r'\d{5}', line).group()
                if not train_date and re.search(r'\d{2}-\d{2}-\d{4}', line):
                    train_date = re.search(r'\d{2}-\d{2}-\d{4}', line).group()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coach Wise"

        title = f"{train_number} {train_date}".strip()
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=100)
        ws.cell(row=1, column=1).value = title
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        current_col = 1
        max_height = 1

        for coach in sorted(grouped.keys()):
            coach_data = grouped[coach]
            # Sort by station_order
            coach_data.sort(key=lambda x: station_order.index(x[0]) if x[0] in station_order else 999)

            col1 = current_col
            ws.merge_cells(start_row=2, start_column=col1, end_row=2, end_column=col1+2)
            ws.cell(row=2, column=col1).value = coach
            ws.cell(row=2, column=col1).alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=col1).fill = PatternFill(start_color="BDD7EE", fill_type="solid")
            ws.cell(row=2, column=col1).font = Font(bold=True)

            headers = [
                f"Station ({hin_headers[0]})",
                f"Berth ({hin_headers[1]})",
                f"Time ({hin_headers[2]})"
            ]
            for i in range(3):
                cell = ws.cell(row=3, column=col1 + i)
                cell.value = headers[i]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(coach_data):
                station_code = row[0]
                hindi_name = hindi_station_map.get(station_code, "")
                ws.cell(row=4 + i, column=col1).value = f"{station_code} {hindi_name}".strip()
                ws.cell(row=4 + i, column=col1 + 1).value = row[1]
                ws.cell(row=4 + i, column=col1 + 2).value = row[2]

            for i in range(3):
                ws.column_dimensions[get_column_letter(col1 + i)].width = 20

            current_col += 4
            max_height = max(max_height, len(coach_data))

        # Auto fit row height and wrap text
        for row in ws.iter_rows(min_row=4, max_row=4 + max_height + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Save to memory
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return send_file(file_stream, as_attachment=True, download_name=f"{train_number}_coachwise.xlsx")

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
